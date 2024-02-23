import openpyxl

def check_changes(file1, file2):
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    changes = {}
    for sheetname in wb1.sheetnames:
        sheet1 = wb1[sheetname]
        sheet2 = wb2[sheetname]
        for row1, row2 in zip(sheet1.iter_rows(values_only=True), sheet2.iter_rows(values_only=True)):
            if row1 != row2:
                changes[sheetname] = True
                break
        else:
            changes[sheetname] = False

    return changes

changes = check_changes('KIIT School Of Law Delegation Sheet.xlsx', 'KIIT School Of Law Delegation Sheet - Copy.xlsx')
for sheet, changed in changes.items():
    print(f"Sheet {sheet} has {'changed' if changed else 'not changed'} since the last version.")